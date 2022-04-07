import openpyxl
import pandas as pd

def calculate_max_rows(sheets, column):
    ''' Вычисление количества непустых строк на странице
    ВНИМАНИЕ на столбец, по которому происходит просмотр макимального количества строк'''
    not_empty_rows = sheets.max_row  # непустые строки
    while not_empty_rows > 1:
        if sheets.cell(row=not_empty_rows, column=column).value == None:  # ----- проверка по строке D -----
            not_empty_rows -= 1
        else:
            break

    return not_empty_rows


#
file_m = 'monitoring.xlsx'
file_s = 'statistics.xlsx'

wb_m = openpyxl.load_workbook(file_m, read_only=False) #, keep_vba=True)  # data_only=True
# ----- проверка на то, что эксель файл закрыт ---
try: wb_m.save(file_m)
except PermissionError as ex:
    print(f" Ошибка. файл {file_m } открыт. Закройте файл")
    print(input('\n Для продолжения - нажмите "Enter"'))
    exit()

wb_s = openpyxl.load_workbook(file_s, read_only=False) #, keep_vba=True)  # data_only=True
# ----- проверка на то, что эксель файл закрыт ---
try: wb_s.save(file_s)
except PermissionError as ex:
    print(f" Ошибка. файл {file_s } открыт. Закройте файл")
    print(input('\n Для продолжения - нажмите "Enter"'))
    exit()


ws_m = wb_m['Лист1']
ws_s = wb_s['Лист1']

error_e = 0
#  ----- определим w - считаем из таблицы ----
w = ws_s.cell(row=2, column=18).value


# ----- вычисляем нижнюю заполненную строку в файле Мониторинг -----
last_row_m = calculate_max_rows(ws_m, 2)
# ----- вычисляем нижнюю заполненную строку в файле Статистика -----
last_row_s = calculate_max_rows(ws_s, 2)

# ---- очистим статистику от прежних занчаний ----
ws_s.delete_rows(3, last_row_s)


# print(input(f'Вывожу результат анализа первой строки (нажмите "Enter")'))

# actual_proc_data = {}
for item_k in range(last_row_m - 1):
    actual_proc_data = {}
    row_ = item_k + 3
    # ---- соберем данные о текущем процессе и запишем его в файл Статистика -----
    actual_proc_data['id'] = ws_m.cell(row=row_ - 1, column=1).value
    actual_proc_data['create_time'] = ws_m.cell(row=row_ - 1, column=2).value
    actual_proc_data['ip_address'] = ws_m.cell(row=row_ - 1, column=3).value
    actual_proc_data['port'] = ws_m.cell(row=row_ - 1, column=4).value
    actual_proc_data['ppid'] = ws_m.cell(row=row_ - 1, column=5).value
    actual_proc_data['count_proc_copy'] = ws_m.cell(row=row_ - 1, column=6).value
    #  ----- рассчитаем и запишем в словарь среднее время процесса ср.кол. копий(для текущего) ----
    # ---- просто перенесем из уже рассчитанных ячеек файла Мониторинг ----
    actual_proc_data['time_means'] = ws_m.cell(row=row_ - 1, column=8).value # time
    actual_proc_data['cpc_means'] = ws_m.cell(row=row_ - 1, column=7).value # cpc

    # ----- Занесем данные полученного словаря в таблицу СТАТИСТИКА -----
    ws_s.cell(row=row_, column=1).value = actual_proc_data['id']
    ws_s.cell(row=row_, column=2).value = actual_proc_data['create_time']
    ws_s.cell(row=row_, column=3).value = actual_proc_data['time_means']

    ws_s.cell(row=row_, column=5).value = actual_proc_data['ip']
    # ----- если первая строка, то запишем текущий IP = предыдущему типа -----
    if row_ == 3: ws_s.cell(row=row_, column=6).value = actual_proc_data['ip']
    else: ws_s.cell(row=row_, column=6).value = ws_s.cell(row=row_ - 1, column=5).value # ----- предыдущий IP

    ws_s.cell(row=row_, column=8).value = actual_proc_data['port']
    # ----- если первая строка, то запишем текущий PORT = предыдущему типа -----
    if row_ == 3: ws_s.cell(row=row_, column=9).value = actual_proc_data['port']
    else: ws_s.cell(row=row_, column=9).value = ws_s.cell(row=row_ - 1, column=8).value # ----- предыдущий PORT

    ws_s.cell(row=row_, column=11).value = actual_proc_data['ppid']
    # ----- если первая строка, то запишем текущий ppid = предыдущему типа -----
    if row_ == 3: ws_s.cell(row=row_, column=12).value = actual_proc_data['ppid']
    else: ws_s.cell(row=row_, column=12).value = ws_s.cell(row=row_ - 1, column=11).value # ----- предыдущий ppid

    ws_s.cell(row=row_, column=14).value = actual_proc_data['count_proc_copy']
    ws_s.cell(row=row_, column=15).value = actual_proc_data['cpc_means']
    # ws_s.delete_rows(row_ + 1)

    # ---- ПРОВЕРКА НА АНОМАЛЬНОСТЬ  процесса ----
    # ---- актуальные данные:
    time_actual = actual_proc_data['create_time']
    ip_actual = actual_proc_data['ip']
    port_actual = actual_proc_data['port']
    ppid_actual = actual_proc_data['ppid']
    cpc_actual = actual_proc_data['count_proc_copy']

    # ----- предыдущие данные ------
    time_m = ws_s.cell(row=row_, column=3).value
    ip_prev = ws_s.cell(row=row_, column=6).value
    port_prev = ws_s.cell(row=row_, column=9).value
    ppid_prev = ws_s.cell(row=row_, column=12).value
    cpc_m = ws_s.cell(row=row_, column=15).value

    # ----- Рассчитакем коэффициенты К -------
    if time_actual == None:
        break
    if time_actual > time_m:  k1 = 20
    else: k1 = 0
    if ip_actual != ip_prev:  k2 = 20
    else: k2 = 0
    if port_actual != port_prev:  k3 = 20
    else: k3 = 0
    if ppid_actual != ppid_prev:  k4 = 20
    else: k4 = 0
    if cpc_actual > cpc_m:  k5 = 20
    else: k5 = 0

    # ----- запишем коэффициенты в таблицу ---
    ws_s.cell(row=row_, column=4).value = k1
    ws_s.cell(row=row_, column=7).value = k2
    ws_s.cell(row=row_, column=10).value = k3
    ws_s.cell(row=row_, column=13).value = k4
    ws_s.cell(row=row_, column=16).value = k5

    error_e += w * (k1 + k2 + k3 + k4 + k5)
    ws_s.cell(row=row_, column=17).value = error_e


wb_s.save(file_s)
wb_s.close()
wb_m.close()
